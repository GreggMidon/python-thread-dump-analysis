'''
Name....: .py
Date....:
Author..: Gregg Midon
Desc....:
Usage...:

'''
import os,sys
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, NamedStyle, numbers, PatternFill
from openpyxl.cell import g

def main():

    # open the template worksheet
    wb = load_workbook('c:\\Development\\scripts\\ash.xlsx')

    # worksheet reference
    ws = wb.get_sheet_by_name('data')

    # progress through working arae
    row = 5
    col = 14
    col_letter = get_col_letter(col)

    ws[] =
    ws[] =
    ws[] =
    ws[] =
    ws[] =
    ws[] =









if __name__ == '__main__':
    main()
